# backend/app/api/routes.py

from fastapi import APIRouter, HTTPException, BackgroundTasks, Depends, UploadFile, File
from fastapi.responses import FileResponse, StreamingResponse
from sqlalchemy.orm import Session
from typing import List, Optional
from pydantic import BaseModel
import pandas as pd
import io
from pathlib import Path
import shutil
import logging
import zipfile
from datetime import datetime
import uuid  # ✅ ADD THIS IMPORT
from sqlalchemy import func  # ✅ ADD THIS IMPORT

from app.database import get_db
from app.schemas import (
    DocumentDetailSchema,
    ProcessingStatsSchema,
    SystemInfoSchema,
    BatchResultSchema,
    UserInfoCreateSchema,
    UserInfoSchema,
    UserTicketCreateSchema,
    UserTicketSchema
)

# Request models
class StartProcessingRequest(BaseModel):
    max_workers: int = 2          # For legacy mode or total workers
    ocr_workers: Optional[int] = None  # For pipeline mode
    llm_workers: Optional[int] = None  # For pipeline mode
    stage2_queue_size: Optional[int] = None  # Bounded queue size for Stage-2
    enable_ocr_multiprocessing: Optional[bool] = None  # Enable OCR multiprocessing
    ocr_page_workers: Optional[int] = None  # OCR page-level workers

from app.models import DocumentDetail, PropertyDetail, BuyerDetail, SellerDetail, BatchSession, UserInfo, UserTicket
from app.utils.file_handler import FileHandler
from app.config import settings

logger = logging.getLogger(__name__)
router = APIRouter()

# Initialize processors based on pipeline mode
if settings.ENABLE_PIPELINE:
    # Version 2: Pipeline mode
    from app.workers.pipeline_processor_v2 import PipelineBatchProcessor
    from app.services.pdf_processor_v2 import PDFProcessorV2

    pipeline_processor = PipelineBatchProcessor(
        max_ocr_workers=settings.MAX_OCR_WORKERS,
        max_llm_workers=settings.MAX_LLM_WORKERS
    )
    pdf_processor_v2 = PDFProcessorV2(batch_processor=pipeline_processor)
    batch_processor = pipeline_processor  # For compatibility
    logger.info("API initialized with Pipeline Processor V2")
else:
    # Version 1: Legacy mode
    from app.workers.batch_processor import BatchProcessor
    from app.services.pdf_processor import PDFProcessor

    batch_processor = BatchProcessor()
    pdf_processor = PDFProcessor(batch_processor=batch_processor)
    logger.info("API initialized with Legacy Batch Processor V1")

# Vision processor (same for both versions)
from app.workers.vision_batch_processor import VisionBatchProcessor
vision_batch_processor = VisionBatchProcessor()

# ==================== UPLOAD ENDPOINTS ====================

@router.post("/upload", response_model=dict)
async def upload_pdfs(files: List[UploadFile] = File(...), db: Session = Depends(get_db)):
    """Upload PDF files to newly_uploaded folder and assign batch ID"""
    try:
        # ✅ GENERATE UNIQUE BATCH ID
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        unique_id = str(uuid.uuid4())[:8]
        batch_id = f"BATCH_{timestamp}_{unique_id}"
        
        uploaded_files = []

        # ✅ GET FIRST DOCUMENT NAME FOR BATCH NAME
        first_document_name = None
        for file in files:
            if file.filename.endswith('.pdf'):
                # Remove .pdf extension and use as batch name
                first_document_name = file.filename.replace('.pdf', '')
                break

        # ✅ CREATE BATCH SESSION RECORD
        try:
            batch_session = BatchSession(
                batch_id=batch_id,
                uploaded_count=len(files),
                batch_name=first_document_name or f"Upload {timestamp}",  # Use first document name
                status='pending'
            )
            db.add(batch_session)
        except Exception as e:
            logger.warning(f"Could not create batch session: {e}")
            # Continue even if batch session creation fails

        for file in files:
            if not file.filename.endswith('.pdf'):
                continue
            
            file_path = settings.NEWLY_UPLOADED_DIR / file.filename
            
            with open(file_path, "wb") as buffer:
                shutil.copyfileobj(file.file, buffer)
            
            uploaded_files.append(file.filename)
            logger.info(f"Uploaded: {file.filename}")
            
            # ✅ CREATE OR UPDATE DOCUMENT RECORD WITH BATCH_ID
            doc_id = FileHandler.extract_document_id(file.filename)
            
            existing_doc = db.query(DocumentDetail).filter(
                DocumentDetail.document_id == doc_id
            ).first()
            
            if not existing_doc:
                new_doc = DocumentDetail(
                    document_id=doc_id,
                    batch_id=batch_id  # ✅ ASSIGN BATCH ID
                )
                db.add(new_doc)
            else:
                existing_doc.batch_id = batch_id  # ✅ UPDATE BATCH ID
        
        db.commit()
        
        return {
            "success": True,
            "uploaded_count": len(uploaded_files),
            "files": uploaded_files,
            "batch_id": batch_id  # ✅ RETURN BATCH ID
        }
    
    except Exception as e:
        db.rollback()
        logger.error(f"Upload error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ✅ DUPLICATE DETECTION ENDPOINT
@router.post("/check-duplicates", response_model=dict)
async def check_duplicates(files: List[UploadFile] = File(...), db: Session = Depends(get_db)):
    """Check if uploaded files are duplicates of existing documents"""
    try:
        from app.utils.duplicate_detector import calculate_file_hash, check_duplicate
        import tempfile
        
        duplicates = []
        unique_files = []
        
        for file in files:
            if not file.filename.endswith('.pdf'):
                continue
            
            # Save to temp file to calculate hash
            with tempfile.NamedTemporaryFile(delete=False, suffix='.pdf') as temp_file:
                content = await file.read()
                temp_file.write(content)
                temp_path = Path(temp_file.name)
            
            try:
                # Calculate hash
                file_hash = calculate_file_hash(temp_path)
                
                # Check for duplicate
                duplicate_info = check_duplicate(file_hash, db)
                
                if duplicate_info:
                    duplicates.append({
                        "filename": file.filename,
                        "hash": file_hash,
                        "existing_document": duplicate_info
                    })
                else:
                    unique_files.append({
                        "filename": file.filename,
                        "hash": file_hash
                    })
                    
            finally:
                # Clean up temp file
                temp_path.unlink(missing_ok=True)
                # Reset file pointer for potential re-upload
                await file.seek(0)
        
        return {
            "success": True,
            "total_files": len(files),
            "duplicates_found": len(duplicates),
            "unique_files": len(unique_files),
            "duplicates": duplicates,
            "unique": unique_files
        }
        
    except Exception as e:
        logger.error(f"Duplicate check error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ✅ ADD NEW ENDPOINT: Get all batches
@router.get("/batches", response_model=List[dict])
async def get_batches(db: Session = Depends(get_db)):
    """Get all batch sessions with document counts"""
    try:
        # Query batch sessions with document counts
        batches = db.query(BatchSession).order_by(BatchSession.created_at.desc()).all()

        result = []
        for batch_session in batches:
            # Count documents in this batch
            doc_count = db.query(DocumentDetail).filter(
                DocumentDetail.batch_id == batch_session.batch_id
            ).count()

            # Use processing_started_at if available, otherwise use created_at
            display_time = batch_session.processing_started_at or batch_session.created_at

            result.append({
                "batch_id": batch_session.batch_id,
                "batch_name": batch_session.batch_name,  # First document name
                "total_count": doc_count,
                "created_at": display_time.isoformat() if display_time else None,
                "status": batch_session.status
            })

        return result

    except Exception as e:
        logger.error(f"Failed to fetch batches: {e}")
        raise HTTPException(status_code=500, detail=f"Failed to fetch batches: {str(e)}")


# ==================== PROCESSING ENDPOINTS ====================

@router.post("/process/start", response_model=dict)
async def start_batch_processing(background_tasks: BackgroundTasks, request: StartProcessingRequest = StartProcessingRequest(), db: Session = Depends(get_db)):
    """Start PDF batch processing with configurable worker count (supports both V1 and V2)"""
    try:
        if batch_processor.is_running:
            raise HTTPException(status_code=400, detail="Batch processing already running")

        pdf_files = FileHandler.get_pdf_files(settings.NEWLY_UPLOADED_DIR)

        if not pdf_files:
            return {
                "success": False,
                "message": "No PDF files found in newly_uploaded folder"
            }

        # ✅ UPDATE BATCH SESSION WITH PROCESSING START TIME
        try:
            # Get the most recent batch session with pending status
            batch_session = db.query(BatchSession).filter(
                BatchSession.status == 'pending'
            ).order_by(BatchSession.created_at.desc()).first()

            if batch_session:
                batch_session.processing_started_at = datetime.utcnow()
                batch_session.status = 'processing'
                db.commit()
                logger.info(f"Updated batch {batch_session.batch_id} processing start time")
        except Exception as e:
            logger.warning(f"Could not update batch session processing start time: {e}")
            db.rollback()

        if settings.ENABLE_PIPELINE:
            # Pipeline Mode (V2): Separate OCR and LLM workers
            ocr_workers = request.ocr_workers if request.ocr_workers is not None else settings.MAX_OCR_WORKERS
            llm_workers = request.llm_workers if request.llm_workers is not None else settings.MAX_LLM_WORKERS

            # Handle new settings (apply runtime overrides)
            if request.stage2_queue_size is not None:
                if request.stage2_queue_size < 1 or request.stage2_queue_size > 10:
                    raise HTTPException(status_code=400, detail="stage2_queue_size must be between 1 and 10")
                settings.STAGE2_QUEUE_SIZE = request.stage2_queue_size

            if request.enable_ocr_multiprocessing is not None:
                settings.ENABLE_OCR_MULTIPROCESSING = request.enable_ocr_multiprocessing

            if request.ocr_page_workers is not None:
                if request.ocr_page_workers < 1 or request.ocr_page_workers > 8:
                    raise HTTPException(status_code=400, detail="ocr_page_workers must be between 1 and 8")
                settings.OCR_PAGE_WORKERS = request.ocr_page_workers

            # Validate worker counts (1-20)
            if ocr_workers < 1 or ocr_workers > 20:
                raise HTTPException(status_code=400, detail="ocr_workers must be between 1 and 20")
            if llm_workers < 1 or llm_workers > 20:
                raise HTTPException(status_code=400, detail="llm_workers must be between 1 and 20")

            # Update pipeline processor
            pipeline_processor.max_ocr_workers = ocr_workers
            pipeline_processor.max_llm_workers = llm_workers

            logger.info(
                f"Starting pipeline processing: {ocr_workers} OCR + {llm_workers} LLM workers, "
                f"Queue size: {settings.STAGE2_QUEUE_SIZE}, "
                f"OCR multiprocessing: {settings.ENABLE_OCR_MULTIPROCESSING}, "
                f"OCR page workers: {settings.OCR_PAGE_WORKERS}"
            )

            # Start pipeline processing
            background_tasks.add_task(
                pipeline_processor.process_batch,
                pdf_files,
                pdf_processor_v2,  # Stage 1 processor
                pdf_processor_v2   # Stage 2 processor
            )

            return {
                "success": True,
                "message": f"Started processing {len(pdf_files)} PDFs with {ocr_workers} OCR + {llm_workers} LLM workers",
                "total_files": len(pdf_files),
                "ocr_workers": ocr_workers,
                "llm_workers": llm_workers,
                "stage2_queue_size": settings.STAGE2_QUEUE_SIZE,
                "enable_ocr_multiprocessing": settings.ENABLE_OCR_MULTIPROCESSING,
                "ocr_page_workers": settings.OCR_PAGE_WORKERS,
                "pipeline_mode": True
            }

        else:
            # Legacy Mode (V1): Single worker pool
            max_workers = request.max_workers

            # Validate max_workers range (1-20)
            if max_workers < 1 or max_workers > 20:
                raise HTTPException(status_code=400, detail="max_workers must be between 1 and 20")

            batch_processor.max_workers = max_workers
            logger.info(f"Starting batch processing with {max_workers} workers")

            # Start processing in background
            background_tasks.add_task(
                batch_processor.process_batch,
                pdf_files,
                pdf_processor.process_single_pdf
            )

            return {
                "success": True,
                "message": f"Started processing {len(pdf_files)} PDFs with {max_workers} workers",
                "total_files": len(pdf_files),
                "max_workers": max_workers,
                "pipeline_mode": False
            }

    except Exception as e:
        logger.error(f"Start processing error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/process/stop", response_model=dict)
async def stop_batch_processing():
    """Stop PDF OCR batch processing (completes current tasks)"""
    try:
        if not batch_processor.is_running:
            return {
                "success": False,
                "message": "No batch processing is running"
            }

        batch_processor.stop()

        # Get stats to calculate how many might be stopped
        stats = batch_processor.get_stats()
        remaining = stats.get("total", 0) - stats.get("processed", 0)

        return {
            "success": True,
            "message": f"Processing stopped. {remaining} PDF(s) remaining in newly uploaded folder.",
            "stopped_count": remaining
        }

    except Exception as e:
        logger.error(f"Stop processing error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/process/stats", response_model=ProcessingStatsSchema)
async def get_processing_stats():
    """Get current processing statistics"""
    return batch_processor.get_stats()

@router.post("/process/toggle-embedded-ocr", response_model=dict)
async def toggle_embedded_ocr(enabled: bool):
    """Toggle embedded OCR mode (PyMuPDF vs Poppler+Tesseract)"""
    try:
        settings.USE_EMBEDDED_OCR = enabled
        
        # Persist to runtime config file
        from app.utils.runtime_config import update_runtime_config
        update_runtime_config("USE_EMBEDDED_OCR", enabled)
        
        ocr_mode = "PyMuPDF (Embedded OCR)" if enabled else "Poppler+Tesseract"
        return {
            "success": True,
            "message": f"Embedded OCR mode {'enabled' if enabled else 'disabled'} and saved",
            "use_embedded_ocr": enabled,
            "ocr_mode": ocr_mode
        }
    except Exception as e:
        logger.error(f"Toggle embedded OCR error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/process/rerun-failed", response_model=dict)
async def rerun_failed_pdfs():
    """Move all PDFs from failed folder back to newly_uploaded folder for reprocessing"""
    try:
        failed_dir = settings.FAILED_DIR
        newly_uploaded_dir = settings.NEWLY_UPLOADED_DIR

        failed_files = list(failed_dir.glob("*.pdf"))

        if not failed_files:
            return {
                "success": False,
                "message": "No failed PDFs to rerun"
            }

        moved_count = 0
        for pdf_file in failed_files:
            dest_path = newly_uploaded_dir / pdf_file.name
            shutil.move(str(pdf_file), str(dest_path))
            moved_count += 1
            logger.info(f"Moved {pdf_file.name} from failed to newly_uploaded")

        return {
            "success": True,
            "message": f"Moved {moved_count} failed PDFs back to newly_uploaded folder",
            "moved_count": moved_count
        }

    except Exception as e:
        logger.error(f"Rerun failed PDFs error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/process/rerun-failed-batch", response_model=dict)
async def rerun_failed_batch_pdfs(batch_id: str):
    """Move PDFs from failed folder back to newly_uploaded folder for a specific batch"""
    try:
        failed_dir = settings.FAILED_DIR
        newly_uploaded_dir = settings.NEWLY_UPLOADED_DIR
        
        # Get all failed files
        failed_files = list(failed_dir.glob("*.pdf"))
        
        # Filter files by batch_id (batch_id is in filename)
        batch_files = [f for f in failed_files if batch_id in f.name]
        
        if not batch_files:
            return {
                "success": False,
                "message": f"No failed PDFs found for batch {batch_id}"
            }
        
        moved_count = 0
        for pdf_file in batch_files:
            dest_path = newly_uploaded_dir / pdf_file.name
            shutil.move(str(pdf_file), str(dest_path))
            moved_count += 1
            logger.info(f"Moved {pdf_file.name} from failed to newly_uploaded (batch: {batch_id})")
        
        return {
            "success": True,
            "message": f"Moved {moved_count} failed PDFs for batch {batch_id}",
            "moved_count": moved_count,
            "batch_id": batch_id
        }
    
    except Exception as e:
        logger.error(f"Rerun failed batch error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/process/download-failed")
async def download_failed_pdfs():
    """Download all failed PDFs as a ZIP file"""
    try:
        failed_dir = settings.FAILED_DIR
        failed_files = list(failed_dir.glob("*.pdf"))

        if not failed_files:
            raise HTTPException(status_code=404, detail="No failed PDFs found")

        # Create ZIP in memory
        zip_buffer = io.BytesIO()
        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zip_file:
            for pdf_file in failed_files:
                zip_file.write(pdf_file, pdf_file.name)
                logger.info(f"Added {pdf_file.name} to ZIP")

        zip_buffer.seek(0)

        filename = f"failed_pdfs_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"

        return StreamingResponse(
            zip_buffer,
            media_type="application/zip",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        logger.error(f"Download failed PDFs error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/process/failed-documents")
async def get_failed_documents(db: Session = Depends(get_db)):
    """Get list of failed documents grouped by batch"""
    try:
        failed_dir = settings.FAILED_DIR
        failed_files = list(failed_dir.glob("*.pdf"))
        
        # Group by batch_id
        batch_failures = {}
        
        for file in failed_files:
            # Extract batch_id from filename (format: {batch_id}_{doc_id}.pdf)
            parts = file.stem.split('_')
            if len(parts) >= 2:
                batch_id = '_'.join(parts[:-1])  # Everything except last part
                doc_id = parts[-1]
                
                if batch_id not in batch_failures:
                    # Get batch info from database
                    batch = db.query(BatchSession).filter(
                        BatchSession.batch_id == batch_id
                    ).first()
                    
                    # Count total documents for this batch
                    total_docs = db.query(DocumentDetail).filter(
                        DocumentDetail.batch_id == batch_id
                    ).count()
                    
                    batch_failures[batch_id] = {
                        "batch_id": batch_id,
                        "batch_name": batch.batch_name if batch else batch_id,
                        "total_processed": total_docs,
                        "failed_count": 0,
                        "failed_files": []
                    }
                
                batch_failures[batch_id]["failed_count"] += 1
                batch_failures[batch_id]["failed_files"].append({
                    "filename": file.name,
                    "document_id": doc_id,
                    "size": file.stat().st_size,
                    "modified": datetime.fromtimestamp(file.stat().st_mtime).isoformat()
                })
        
        # Convert to list and sort by failed_count desc
        result = list(batch_failures.values())
        result.sort(key=lambda x: x["failed_count"], reverse=True)
        
        return {
            "success": True,
            "total_failed_files": len(failed_files),
            "batch_count": len(result),
            "batches": result
        }
    
    except Exception as e:
        logger.error(f"Get failed documents error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

# ==================== VISION PROCESSING ENDPOINTS ====================

@router.post("/vision/start", response_model=dict)
async def start_vision_processing(background_tasks: BackgroundTasks):
    """Start vision model batch processing for registration fee tables"""
    try:
        if vision_batch_processor.is_running:
            raise HTTPException(status_code=400, detail="Vision processing already running")
        
        # Start processing in background
        background_tasks.add_task(vision_batch_processor.process_batch)
        
        return {
            "success": True,
            "message": "Started vision processing for registration fee extraction"
        }
    
    except Exception as e:
        logger.error(f"Start vision processing error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/vision/stop", response_model=dict)
async def stop_vision_processing():
    """Stop vision batch processing"""
    try:
        if not vision_batch_processor.is_running:
            return {
                "success": False,
                "message": "No vision processing is running"
            }

        vision_batch_processor.stop()

        # Get stats
        stats = vision_batch_processor.get_stats()
        remaining = stats.get("total", 0) - stats.get("processed", 0)

        return {
            "success": True,
            "message": f"Vision processing stopped. {remaining} image(s) remaining in left over reg fee folder.",
            "stopped_count": remaining
        }

    except Exception as e:
        logger.error(f"Stop vision processing error: {e}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/vision/stats", response_model=ProcessingStatsSchema)
async def get_vision_stats():
    """Get vision processing statistics"""
    return vision_batch_processor.get_stats()

# ==================== DATA RETRIEVAL ENDPOINTS ====================

@router.get("/documents", response_model=List[DocumentDetailSchema])
async def get_all_documents(
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db)
):
    """Get all documents with pagination"""
    documents = db.query(DocumentDetail).offset(skip).limit(limit).all()
    return documents

@router.get("/documents/{document_id}", response_model=DocumentDetailSchema)
async def get_document(document_id: str, db: Session = Depends(get_db)):
    """Get specific document by ID"""
    document = db.query(DocumentDetail).filter(
        DocumentDetail.document_id == document_id
    ).first()
    
    if not document:
        raise HTTPException(status_code=404, detail="Document not found")
    
    return document

def format_number(value):
    """Format number to remove unnecessary decimals"""
    if value is None:
        return None
    # If it's a whole number, convert to int, otherwise keep as float
    if isinstance(value, (int, float)) and value % 1 == 0:
        return int(value)
    return value

@router.get("/export/excel")
async def export_to_excel(
    start_index: int = 0,
    end_index: Optional[int] = None,
    batch_ids: Optional[str] = None,
    batch_names: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    download_type: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Export data to Excel file (optionally filtered by batch IDs)"""
    try:
        from sqlalchemy.orm import joinedload
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # ✅ ADD EAGER LOADING FOR RELATIONSHIPS
        query = db.query(DocumentDetail).options(
            joinedload(DocumentDetail.property_details),
            joinedload(DocumentDetail.buyers),
            joinedload(DocumentDetail.sellers)
        )

        # FILTER BY BATCH IDS IF PROVIDED
        if batch_ids:
            batch_list = batch_ids.split(',')
            query = query.filter(DocumentDetail.batch_id.in_(batch_list))

        query = query.offset(start_index)

        if end_index:
            query = query.limit(end_index - start_index)

        documents = query.all()

        # Prepare data for Excel with new format
        rows = []
        serial_number = 1

        for doc in documents:
            # Prepare property data
            prop = doc.property_details

            # Create Schedule C address with property name
            sched_c_address = ''
            if prop:
                parts = []
                if prop.schedule_c_property_address:
                    parts.append(prop.schedule_c_property_address)
                if prop.schedule_c_property_name:
                    parts.append(prop.schedule_c_property_name)
                sched_c_address = ', '.join(parts) if parts else None

            base_row = {
                "SL_NO": serial_number,
                "Document_ID": doc.document_id,
                "Schedule_B_Area_sqft": format_number(prop.schedule_b_area) if prop else None,
                "Schedule_C_Area_sqft": format_number(prop.schedule_c_property_area) if prop else None,
                "Schedule_C_Address_Name": sched_c_address,
                "Property_Pincode": prop.pincode if prop else None,
                "Property_State": prop.state if prop else None,
                "Sale_Consideration": format_number(prop.sale_consideration) if prop else None,
                "Stamp_Duty_Fee": format_number(prop.stamp_duty_fee) if prop else None,
                "Registration_Fee": format_number(prop.registration_fee) if prop else None,
                "Guidance_Value": format_number(prop.guidance_value) if prop else None,
                "Cash_Payment": prop.paid_in_cash_mode if prop else None,
                "Transaction_Date": doc.transaction_date,
                "Registration_Office": doc.registration_office
            }

            # Add sellers first (USER_TYPE = S)
            for seller in doc.sellers:
                row = base_row.copy()
                row.update({
                    "USER_TYPE": "S",
                    "Name": seller.name,
                    "Gender": seller.gender,
                    "Aadhaar": seller.aadhaar_number,
                    "PAN": seller.pan_card_number,
                    "Address": seller.address,
                    "Pincode": seller.pincode,
                    "State": seller.state,
                    "Phone": seller.phone_number,
                    "Secondary_Phone": seller.secondary_phone_number,
                    "Email": seller.email,
                    "Property_Share": seller.property_share
                })
                rows.append(row)

            # Add buyers (USER_TYPE = B)
            for buyer in doc.buyers:
                row = base_row.copy()
                row.update({
                    "USER_TYPE": "B",
                    "Name": buyer.name,
                    "Gender": buyer.gender,
                    "Aadhaar": buyer.aadhaar_number,
                    "PAN": buyer.pan_card_number,
                    "Address": buyer.address,
                    "Pincode": buyer.pincode,
                    "State": buyer.state,
                    "Phone": buyer.phone_number,
                    "Secondary_Phone": buyer.secondary_phone_number,
                    "Email": buyer.email,
                    "Property_Share": None
                })
                rows.append(row)

            serial_number += 1

        # Check if we have any rows
        if not rows:
            raise HTTPException(status_code=404, detail="No documents found for the selected criteria")

        df = pd.DataFrame(rows)

        # Reorder columns to match desired format
        column_order = [
            "SL_NO", "USER_TYPE", "Document_ID",
            "Schedule_B_Area_sqft", "Schedule_C_Area_sqft", "Schedule_C_Address_Name",
            "Property_Pincode", "Property_State",
            "Sale_Consideration", "Stamp_Duty_Fee", "Registration_Fee", "Guidance_Value",
            "Cash_Payment", "Transaction_Date", "Registration_Office",
            "Name", "Gender", "Aadhaar", "PAN", "Address", "Pincode", "State",
            "Phone", "Secondary_Phone", "Email", "Property_Share"
        ]
        df = df[column_order]

        # Create Excel file in memory with styling
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Sale_Deeds')

            # Get the worksheet to apply styling
            worksheet = writer.sheets['Sale_Deeds']

            # Define header styles
            header_font = Font(bold=True, size=11, color='FFFFFF')
            header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
            header_alignment = Alignment(vertical='top', horizontal='center', wrap_text=True)
            border = Border(
                left=Side(style='thin', color='000000'),
                right=Side(style='thin', color='000000'),
                top=Side(style='thin', color='000000'),
                bottom=Side(style='thin', color='000000')
            )

            # Apply styles to header row
            for cell in worksheet[1]:
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = border

            # Set row height for header
            worksheet.row_dimensions[1].height = 25

            # Set column widths
            column_widths = {
                'A': 8, 'B': 10, 'C': 22, 'D': 18, 'E': 18, 'F': 45,
                'G': 14, 'H': 14, 'I': 16, 'J': 13, 'K': 15, 'L': 15,
                'M': 35, 'N': 14, 'O': 20, 'P': 28, 'Q': 10, 'R': 15,
                'S': 12, 'T': 35, 'U': 10, 'V': 14, 'W': 14, 'X': 14,
                'Y': 26, 'Z': 14
            }

            for col, width in column_widths.items():
                worksheet.column_dimensions[col].width = width

            # Apply borders and alignment to data rows (only if there are data rows)
            if worksheet.max_row > 1:
                for row_idx, row in enumerate(worksheet.iter_rows(min_row=2, max_row=worksheet.max_row), start=2):
                    for cell in row:
                        cell.border = Border(
                            left=Side(style='thin', color='D0D0D0'),
                            right=Side(style='thin', color='D0D0D0'),
                            top=Side(style='thin', color='D0D0D0'),
                            bottom=Side(style='thin', color='D0D0D0')
                        )
                        cell.alignment = Alignment(vertical='top', horizontal='left', wrap_text=True)
                        cell.font = Font(size=10)
                    # Set row height once per row
                    worksheet.row_dimensions[row_idx].height = 20

        output.seek(0)

        # Generate filename based on download type
        filename = "sale_deeds_export.xlsx"

        if download_type == "all":
            filename = "whole_data.xlsx"
        elif download_type == "batch" and batch_names:
            # Use first batch name if single batch, otherwise use "multiple_batches"
            names_list = batch_names.split(',')
            if len(names_list) == 1:
                # Sanitize batch name for filename
                safe_name = names_list[0].replace('/', '_').replace('\\', '_').replace(' ', '_')
                filename = f"{safe_name}.xlsx"
            else:
                filename = f"multiple_batches_{len(names_list)}.xlsx"
        elif download_type == "dateRange" and start_date and end_date:
            filename = f"{start_date}_to_{end_date}.xlsx"
        elif download_type == "dateRange" and start_date:
            filename = f"from_{start_date}.xlsx"
        elif download_type == "dateRange" and end_date:
            filename = f"to_{end_date}.xlsx"

        return StreamingResponse(
            output,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )

    except Exception as e:
        import traceback
        error_detail = f"{str(e)}\n{traceback.format_exc()}"
        logger.error(f"Export error: {error_detail}")
        raise HTTPException(status_code=500, detail=str(e))

@router.get("/export/csv")
async def export_to_csv(
    batch_ids: Optional[str] = None,
    batch_names: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    download_type: Optional[str] = None,
    db: Session = Depends(get_db)
):
    """Export data to CSV file with 42-column format (optionally filtered by batch IDs or date range)"""
    try:
        from app.services.csv_export_service import generate_csv_export
        
        # Generate CSV
        csv_output = generate_csv_export(
            db=db,
            batch_ids=batch_ids,
            start_date=start_date,
            end_date=end_date
        )
        
        # Generate filename based on download type
        filename = "sale_deeds_export.csv"
        
        if download_type == "all":
            filename = "whole_data.csv"
        elif download_type == "batch" and batch_names:
            # Use first batch name if single batch, otherwise use "multiple_batches"
            names_list = batch_names.split(',')
            if len(names_list) == 1:
                # Sanitize batch name for filename
                safe_name = names_list[0].replace('/', '_').replace('\\', '_').replace(' ', '_')
                filename = f"{safe_name}.csv"
            else:
                filename = f"multiple_batches_{len(names_list)}.csv"
        elif download_type == "dateRange" and start_date and end_date:
            filename = f"{start_date}_to_{end_date}.csv"
        elif download_type == "dateRange" and start_date:
            filename = f"from_{start_date}.csv"
        elif download_type == "dateRange" and end_date:
            filename = f"to_{end_date}.csv"
        
        return StreamingResponse(
            iter([csv_output.getvalue()]),
            media_type="text/csv",
            headers={"Content-Disposition": f"attachment; filename={filename}"}
        )
    
    except Exception as e:
        import traceback
        error_detail = f"{str(e)}\n{traceback.format_exc()}"
        logger.error(f"CSV Export error: {error_detail}")
        raise HTTPException(status_code=500, detail=str(e))

# ==================== SYSTEM INFO ENDPOINTS ====================

@router.get("/system/info", response_model=SystemInfoSchema)
async def get_system_info():
    """Get system information and health status"""
    import torch
    import subprocess
    
    # Check CUDA
    cuda_available = torch.cuda.is_available()
    cuda_count = torch.cuda.device_count() if cuda_available else 0
    
    # Check Poppler
    try:
        subprocess.run(["pdfinfo", "-v"], capture_output=True, timeout=5)
        poppler_available = True
    except:
        poppler_available = False
    
    # Check Tesseract
    try:
        subprocess.run(["tesseract", "--version"], capture_output=True, timeout=5)
        tesseract_available = True
    except:
        tesseract_available = False
    
    # Check Ollama (use correct processor based on pipeline mode)
    if settings.ENABLE_PIPELINE:
        ollama_connected = pdf_processor_v2.llm_service.check_connection()
    else:
        ollama_connected = pdf_processor.llm_service.check_connection()
    
    # Check YOLO model
    yolo_model_loaded = settings.YOLO_MODEL_PATH.exists()
    
    return {
        "cuda_available": cuda_available,
        "cuda_device_count": cuda_count,
        "poppler_available": poppler_available,
        "tesseract_available": tesseract_available,
        "ollama_connected": ollama_connected,
        "yolo_model_loaded": yolo_model_loaded
    }

@router.get("/system/config", response_model=dict)
async def get_system_config():
    """Get current pipeline and OCR configuration settings"""
    return {
        "enable_pipeline": settings.ENABLE_PIPELINE,
        "max_ocr_workers": settings.MAX_OCR_WORKERS,
        "max_llm_workers": settings.MAX_LLM_WORKERS,
        "stage2_queue_size": settings.STAGE2_QUEUE_SIZE,
        "enable_ocr_multiprocessing": settings.ENABLE_OCR_MULTIPROCESSING,
        "ocr_page_workers": settings.OCR_PAGE_WORKERS,
        "max_workers": settings.MAX_WORKERS,  # Legacy mode
        "llm_backend": settings.LLM_BACKEND,
        "tesseract_lang": settings.TESSERACT_LANG,
        "poppler_dpi": settings.POPPLER_DPI,
        "use_embedded_ocr": settings.USE_EMBEDDED_OCR
    }

@router.get("/system/folders", response_model=dict)
async def get_folder_stats():
    """Get file counts in each folder"""
    return {
        "newly_uploaded": len(list(settings.NEWLY_UPLOADED_DIR.glob("*.pdf"))),
        "processed": len(list(settings.PROCESSED_DIR.glob("*.pdf"))),
        "failed": len(list(settings.FAILED_DIR.glob("*.pdf"))),
        "left_over_reg_fee": len(list(settings.LEFT_OVER_REG_FEE_DIR.glob("*.png"))) +
                             len(list(settings.LEFT_OVER_REG_FEE_DIR.glob("*.jpg")))
    }


# ===== USER INFO ENDPOINTS =====

@router.post("/user-info", response_model=UserInfoSchema)
async def create_user_info(user_info: UserInfoCreateSchema, db: Session = Depends(get_db)):
    """Create a new user info entry before uploading PDFs"""
    try:
        new_user_info = UserInfo(
            user_name=user_info.user_name,
            number_of_files=user_info.number_of_files,
            file_region=user_info.file_region,
            batch_id=user_info.batch_id
        )
        db.add(new_user_info)
        db.commit()
        db.refresh(new_user_info)

        logger.info(f"User info created: {user_info.user_name} - {user_info.number_of_files} files from {user_info.file_region}")
        return new_user_info
    except Exception as e:
        logger.error(f"Error creating user info: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/user-info", response_model=List[UserInfoSchema])
async def get_all_user_info(db: Session = Depends(get_db)):
    """Get all user info entries"""
    try:
        user_infos = db.query(UserInfo).order_by(UserInfo.created_at.desc()).all()
        return user_infos
    except Exception as e:
        logger.error(f"Error fetching user info: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.patch("/user-info/update-batch")
async def update_user_info_batch(
    batch_update: dict,
    db: Session = Depends(get_db)
):
    """Update the most recent user info entry with batch_id"""
    try:
        batch_id = batch_update.get("batch_id")
        user_name = batch_update.get("user_name")

        if not batch_id or not user_name:
            raise HTTPException(status_code=400, detail="batch_id and user_name are required")

        # Find the most recent user info entry for this user with null batch_id
        user_info = db.query(UserInfo).filter(
            UserInfo.user_name == user_name,
            UserInfo.batch_id == None
        ).order_by(UserInfo.created_at.desc()).first()

        if not user_info:
            raise HTTPException(status_code=404, detail="User info entry not found")

        user_info.batch_id = batch_id
        db.commit()
        db.refresh(user_info)

        logger.info(f"Updated user info {user_info.id} with batch_id: {batch_id}")
        return {"success": True, "message": "User info updated with batch_id"}

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating user info batch: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# ===== USER TICKETS ENDPOINTS =====

@router.post("/tickets", response_model=UserTicketSchema)
async def create_ticket(ticket: UserTicketCreateSchema, db: Session = Depends(get_db)):
    """Create a new support ticket"""
    try:
        new_ticket = UserTicket(
            user_name=ticket.user_name,
            batch_id=ticket.batch_id,
            error_type=ticket.error_type,
            description=ticket.description
        )
        db.add(new_ticket)
        db.commit()
        db.refresh(new_ticket)

        logger.info(f"Ticket created: {ticket.user_name} - {ticket.error_type}")
        
        # ✅ SEND EMAIL NOTIFICATION TO DEVELOPER
        try:
            from app.services.email_service import email_service
            
            ticket_data = {
                "ticket_id": new_ticket.id,
                "user_name": new_ticket.user_name,
                "error_type": new_ticket.error_type,
                "error_description": new_ticket.description,
                "batch_id": new_ticket.batch_id or "N/A",
                "created_at": new_ticket.created_at
            }
            
            # Send email asynchronously (non-blocking)
            email_sent = email_service.send_support_ticket_notification(ticket_data)
            
            if email_sent:
                logger.info(f"Email notification sent for ticket #{new_ticket.id}")
            else:
                logger.warning(f"Failed to send email notification for ticket #{new_ticket.id}")
                
        except Exception as email_error:
            # Don't fail ticket creation if email fails
            logger.error(f"Email notification error: {email_error}")
        
        return new_ticket
    except Exception as e:
        logger.error(f"Error creating ticket: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/tickets", response_model=List[UserTicketSchema])
async def get_all_tickets(db: Session = Depends(get_db)):
    """Get all support tickets"""
    try:
        tickets = db.query(UserTicket).order_by(UserTicket.created_at.desc()).all()
        return tickets
    except Exception as e:
        logger.error(f"Error fetching tickets: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/tickets/{ticket_id}", response_model=UserTicketSchema)
async def get_ticket(ticket_id: int, db: Session = Depends(get_db)):
    """Get a specific ticket by ID"""
    try:
        ticket = db.query(UserTicket).filter(UserTicket.id == ticket_id).first()
        if not ticket:
            raise HTTPException(status_code=404, detail="Ticket not found")
        return ticket
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching ticket: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.patch("/tickets/{ticket_id}/status", response_model=UserTicketSchema)
async def update_ticket_status(
    ticket_id: int,
    status_update: dict,
    db: Session = Depends(get_db)
):
    """Update ticket status"""
    try:
        ticket = db.query(UserTicket).filter(UserTicket.id == ticket_id).first()
        if not ticket:
            raise HTTPException(status_code=404, detail="Ticket not found")

        new_status = status_update.get("status")
        if not new_status:
            raise HTTPException(status_code=400, detail="Status is required")

        # Validate status
        valid_statuses = ["open", "in_progress", "resolved", "closed"]
        if new_status not in valid_statuses:
            raise HTTPException(status_code=400, detail=f"Invalid status. Must be one of: {', '.join(valid_statuses)}")

        ticket.status = new_status

        # Set resolved_at timestamp when status is resolved or closed
        if new_status in ["resolved", "closed"] and not ticket.resolved_at:
            from datetime import datetime
            ticket.resolved_at = datetime.utcnow()

        db.commit()
        db.refresh(ticket)

        logger.info(f"Ticket {ticket_id} status updated to: {new_status}")
        return ticket

    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating ticket status: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))

# ===== NOTIFICATIONS ENDPOINTS =====

from app.models import Notification
from app.services.notification_service import notification_service

@router.get("/notifications", response_model=List[dict])
async def get_notifications(
    limit: int = 50,
    unread_only: bool = False,
    db: Session = Depends(get_db)
):
    """Get all notifications"""
    try:
        notifications = notification_service.get_all_notifications(
            db=db,
            limit=limit,
            unread_only=unread_only
        )
        
        return [{
            "id": n.id,
            "title": n.title,
            "message": n.message,
            "notification_type": n.notification_type,
            "related_id": n.related_id,
            "related_type": n.related_type,
            "is_read": n.is_read,
            "created_at": n.created_at.isoformat() if n.created_at else None,
            "read_at": n.read_at.isoformat() if n.read_at else None
        } for n in notifications]
        
    except Exception as e:
        logger.error(f"Error fetching notifications: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/notifications/unread-count", response_model=dict)
async def get_unread_count(db: Session = Depends(get_db)):
    """Get count of unread notifications"""
    try:
        count = notification_service.get_unread_count(db)
        return {"unread_count": count}
    except Exception as e:
        logger.error(f"Error getting unread count: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.patch("/notifications/{notification_id}/read", response_model=dict)
async def mark_notification_read(
    notification_id: int,
    db: Session = Depends(get_db)
):
    """Mark notification as read"""
    try:
        success = notification_service.mark_as_read(db, notification_id)
        if success:
            return {"success": True, "message": "Notification marked as read"}
        else:
            raise HTTPException(status_code=404, detail="Notification not found")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error marking notification as read: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.patch("/notifications/mark-all-read", response_model=dict)
async def mark_all_read(db: Session = Depends(get_db)):
    """Mark all notifications as read"""
    try:
        count = notification_service.mark_all_as_read(db)
        return {
            "success": True,
            "message": f"Marked {count} notifications as read",
            "count": count
        }
    except Exception as e:
        logger.error(f"Error marking all as read: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.delete("/notifications/{notification_id}", response_model=dict)
async def delete_notification(
    notification_id: int,
    db: Session = Depends(get_db)
):
    """Delete a notification"""
    try:
        success = notification_service.delete_notification(db, notification_id)
        if success:
            return {"success": True, "message": "Notification deleted"}
        else:
            raise HTTPException(status_code=404, detail="Notification not found")
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error deleting notification: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ===== DOCUMENT MANAGEMENT ENDPOINTS =====

from app.schemas.document_schemas import (
    DocumentDetailResponse,
    DocumentUpdateRequest,
    DocumentUpdateResponse,
    PropertyDetailSchema,
    SellerDetailSchema,
    BuyerDetailSchema,
    ConfirmingPartyDetailSchema
)
from app.models import PropertyDetail, SellerDetail, BuyerDetail, ConfirmingPartyDetail
from datetime import datetime

@router.get("/documents/{document_id}", response_model=DocumentDetailResponse)
async def get_document(document_id: str, db: Session = Depends(get_db)):
    """Get a single document with all related data"""
    try:
        doc = db.query(DocumentDetail).filter(
            DocumentDetail.document_id == document_id
        ).first()
        
        if not doc:
            raise HTTPException(status_code=404, detail="Document not found")
        
        return doc
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error fetching document: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.patch("/documents/{document_id}", response_model=DocumentUpdateResponse)
async def update_document(
    document_id: str,
    update_data: DocumentUpdateRequest,
    db: Session = Depends(get_db)
):
    """Update document data (manual correction)"""
    try:
        # Get document
        doc = db.query(DocumentDetail).filter(
            DocumentDetail.document_id == document_id
        ).first()
        
        if not doc:
            raise HTTPException(status_code=404, detail="Document not found")
        
        updated_fields = []
        
        # Update document fields
        if update_data.transaction_date is not None:
            from datetime import datetime
            doc.transaction_date = datetime.fromisoformat(update_data.transaction_date).date()
            updated_fields.append('transaction_date')
        
        if update_data.registration_office is not None:
            doc.registration_office = update_data.registration_office
            updated_fields.append('registration_office')
        
        # Update property details
        if update_data.property is not None:
            prop = doc.property_details
            if not prop:
                prop = PropertyDetail(document_id=document_id)
                db.add(prop)
            
            prop_data = update_data.property.model_dump(exclude_unset=True)
            for key, value in prop_data.items():
                setattr(prop, key, value)
                updated_fields.append(f'property.{key}')
        
        # Update sellers
        if update_data.sellers is not None:
            # Delete existing sellers
            db.query(SellerDetail).filter(
                SellerDetail.document_id == document_id
            ).delete()
            
            # Add new sellers
            for seller_data in update_data.sellers:
                seller = SellerDetail(
                    document_id=document_id,
                    **seller_data.model_dump(exclude_unset=True)
                )
                db.add(seller)
            updated_fields.append('sellers')
        
        # Update buyers
        if update_data.buyers is not None:
            # Delete existing buyers
            db.query(BuyerDetail).filter(
                BuyerDetail.document_id == document_id
            ).delete()
            
            # Add new buyers
            for buyer_data in update_data.buyers:
                buyer = BuyerDetail(
                    document_id=document_id,
                    **buyer_data.model_dump(exclude_unset=True)
                )
                db.add(buyer)
            updated_fields.append('buyers')
        
        # Update confirming parties
        if update_data.confirming_parties is not None:
            # Delete existing confirming parties
            db.query(ConfirmingPartyDetail).filter(
                ConfirmingPartyDetail.document_id == document_id
            ).delete()
            
            # Add new confirming parties
            for party_data in update_data.confirming_parties:
                party = ConfirmingPartyDetail(
                    document_id=document_id,
                    **party_data.model_dump(exclude_unset=True)
                )
                db.add(party)
            updated_fields.append('confirming_parties')
        
        # Update timestamp
        doc.updated_at = datetime.utcnow()
        
        # Commit changes
        db.commit()
        
        logger.info(f"Document {document_id} updated by {update_data.updated_by or 'unknown'}: {updated_fields}")
        
        return DocumentUpdateResponse(
            success=True,
            message=f"Document updated successfully",
            document_id=document_id,
            updated_fields=updated_fields,
            timestamp=datetime.utcnow()
        )
        
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error updating document: {e}")
        db.rollback()
        raise HTTPException(status_code=500, detail=str(e))


# ===== ADVANCED SEARCH ENDPOINT =====

@router.post("/documents/search", response_model=List[DocumentDetailResponse])
async def search_documents(
    search_query: Optional[str] = None,
    batch_id: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    registration_office: Optional[str] = None,
    min_sale_consideration: Optional[float] = None,
    max_sale_consideration: Optional[float] = None,
    state: Optional[str] = None,
    pincode: Optional[str] = None,
    skip: int = 0,
    limit: int = 100,
    db: Session = Depends(get_db)
):
    """Advanced search with multiple filters"""
    try:
        # Start with base query
        query = db.query(DocumentDetail)
        
        # Apply filters
        if batch_id:
            query = query.filter(DocumentDetail.batch_id == batch_id)
        
        if start_date:
            from datetime import datetime
            start = datetime.fromisoformat(start_date).date()
            query = query.filter(DocumentDetail.transaction_date >= start)
        
        if end_date:
            from datetime import datetime
            end = datetime.fromisoformat(end_date).date()
            query = query.filter(DocumentDetail.transaction_date <= end)
        
        if registration_office:
            query = query.filter(
                DocumentDetail.registration_office.ilike(f"%{registration_office}%")
            )
        
        # Join with property details for property-based filters
        if any([min_sale_consideration, max_sale_consideration, state, pincode]):
            query = query.join(PropertyDetail, PropertyDetail.document_id == DocumentDetail.document_id)
            
            if state:
                query = query.filter(PropertyDetail.state.ilike(f"%{state}%"))
            
            if pincode:
                query = query.filter(PropertyDetail.pincode == pincode)
        
        # Text search across multiple fields
        if search_query:
            search_term = f"%{search_query}%"
            query = query.outerjoin(PropertyDetail).outerjoin(SellerDetail).outerjoin(BuyerDetail)
            query = query.filter(
                or_(
                    DocumentDetail.document_id.ilike(search_term),
                    DocumentDetail.registration_office.ilike(search_term),
                    PropertyDetail.schedule_c_property_address.ilike(search_term),
                    PropertyDetail.state.ilike(search_term),
                    PropertyDetail.pincode.ilike(search_term),
                    SellerDetail.name.ilike(search_term),
                    SellerDetail.aadhaar.ilike(search_term),
                    SellerDetail.pan.ilike(search_term),
                    BuyerDetail.name.ilike(search_term),
                    BuyerDetail.aadhaar.ilike(search_term),
                    BuyerDetail.pan.ilike(search_term)
                )
            )
        
        # Apply pagination
        query = query.offset(skip).limit(limit)
        
        # Execute query
        documents = query.all()
        
        logger.info(f"Search returned {len(documents)} documents")
        return documents
        
    except Exception as e:
        logger.error(f"Search error: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ===== AUDIT LOG ENDPOINTS =====

from app.models import AuditLog
from app.services.audit_service import audit_service

@router.get("/audit-logs", response_model=List[dict])
async def get_audit_logs(
    user_email: Optional[str] = None,
    action_type: Optional[str] = None,
    entity_type: Optional[str] = None,
    start_date: Optional[str] = None,
    end_date: Optional[str] = None,
    limit: int = 100,
    db: Session = Depends(get_db)
):
    """Get audit logs with filters"""
    try:
        from datetime import datetime
        
        start = datetime.fromisoformat(start_date) if start_date else None
        end = datetime.fromisoformat(end_date) if end_date else None
        
        logs = audit_service.get_audit_logs(
            db=db,
            user_email=user_email,
            action_type=action_type,
            entity_type=entity_type,
            start_date=start,
            end_date=end,
            limit=limit
        )
        
        return [{
            "id": log.id,
            "user_email": log.user_email,
            "action_type": log.action_type,
            "entity_type": log.entity_type,
            "entity_id": log.entity_id,
            "action_details": log.action_details,
            "ip_address": log.ip_address,
            "timestamp": log.timestamp.isoformat() if log.timestamp else None,
            "success": log.success,
            "error_message": log.error_message
        } for log in logs]
        
    except Exception as e:
        logger.error(f"Error fetching audit logs: {e}")
        raise HTTPException(status_code=500, detail=str(e))


@router.get("/audit-logs/user/{user_email}/summary", response_model=dict)
async def get_user_activity_summary(
    user_email: str,
    days: int = 30,
    db: Session = Depends(get_db)
):
    """Get user activity summary"""
    try:
        summary = audit_service.get_user_activity_summary(
            db=db,
            user_email=user_email,
            days=days
        )
        return summary
    except Exception as e:
        logger.error(f"Error getting user summary: {e}")
        raise HTTPException(status_code=500, detail=str(e))


# ===== DATA QUALITY DASHBOARD ENDPOINT =====

@router.get("/data-quality/metrics", response_model=dict)
async def get_data_quality_metrics(db: Session = Depends(get_db)):
    """Get data quality metrics for dashboard"""
    try:
        from sqlalchemy import func, case
        
        # Total documents
        total_docs = db.query(func.count(DocumentDetail.document_id)).scalar()
        
        # Documents with property details
        docs_with_property = db.query(func.count(PropertyDetail.document_id)).scalar()
        
        # Field completion rates
        property_fields = [
            'schedule_b_area', 'schedule_c_property_name', 'schedule_c_property_address',
            'sale_consideration', 'stamp_duty_fee', 'registration_fee', 'pincode', 'state'
        ]
        
        field_completion = {}
        for field in property_fields:
            count = db.query(func.count(getattr(PropertyDetail, field))).filter(
                getattr(PropertyDetail, field).isnot(None),
                getattr(PropertyDetail, field) != ''
            ).scalar()
            field_completion[field] = {
                'filled': count,
                'total': docs_with_property,
                'completion_rate': round((count / docs_with_property * 100) if docs_with_property > 0 else 0, 2)
            }
        
        # Seller/Buyer counts
        total_sellers = db.query(func.count(SellerDetail.id)).scalar()
        total_buyers = db.query(func.count(BuyerDetail.id)).scalar()
        
        # Batch statistics
        batch_stats = db.query(
            func.count(BatchSession.batch_id).label('total_batches'),
            func.sum(case((BatchSession.status == 'completed', 1), else_=0)).label('completed'),
            func.sum(case((BatchSession.status == 'processing', 1), else_=0)).label('processing'),
            func.sum(case((BatchSession.status == 'pending', 1), else_=0)).label('pending')
        ).first()
        
        # Recent activity (last 7 days)
        from datetime import datetime, timedelta
        week_ago = datetime.utcnow() - timedelta(days=7)
        recent_docs = db.query(func.count(DocumentDetail.document_id)).filter(
            DocumentDetail.created_at >= week_ago
        ).scalar()
        
        return {
            'overview': {
                'total_documents': total_docs,
                'documents_with_property': docs_with_property,
                'total_sellers': total_sellers,
                'total_buyers': total_buyers,
                'recent_documents_7d': recent_docs
            },
            'field_completion': field_completion,
            'batch_statistics': {
                'total_batches': batch_stats.total_batches or 0,
                'completed': batch_stats.completed or 0,
                'processing': batch_stats.processing or 0,
                'pending': batch_stats.pending or 0
            },
            'quality_score': round(
                sum(f['completion_rate'] for f in field_completion.values()) / len(field_completion)
                if field_completion else 0, 2
            )
        }
        
    except Exception as e:
        logger.error(f"Error getting data quality metrics: {e}")
        raise HTTPException(status_code=500, detail=str(e))
